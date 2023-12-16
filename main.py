import os
from openai import OpenAI
import openpyxl
import time

client = OpenAI(
    api_key="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"
)

# read every first column
wb = openpyxl.load_workbook('tape.xlsx')
sheet = wb['Sheet1']

for i in range(1, sheet.max_row + 1):
    chat_completion = client.chat.completions.create(
        messages=[
            {
                "role" : "user",
                "content": "Movie Review:" + sheet.cell(row=i, column=1).value + "Question: What is the review type of the text provided as Movie Review (positive or negative? Give explanation for your reasoning. Give your answer in the format in which the first line is either 1 (positive) or 0 (negative), then with 2 lines spaces in b/w give your reasoning",
            }
        ],
        model="gpt-3.5-turbo",
    )
    if i % 3 == 0 :
        time.sleep(60)
    # write to second column
    print(chat_completion.choices[0].message.content)
    sheet.cell(row=i, column=2).value = chat_completion.choices[0].message.content[0]
    sheet.cell(row=i, column=3).value = chat_completion.choices[0].message.content
wb.save('tape.xlsx')
    

